import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from datetime import datetime, timedelta
import csv, os
from PIL import Image, ImageTk # Assumindo que PIL está instalado
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill 
from openpyxl.utils import get_column_letter

CSV_FILE = "carrinhos.csv"
SENHA_MESTRE = "1234" 

if not os.path.exists(CSV_FILE):
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow(["codigo", "data_ultima_limpeza", "historico", "periodicidade_dias"])
else:
    with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        if "periodicidade_dias" not in header:
            messagebox.showinfo("Atualização do CSV",
                                "Detectamos uma versão antiga do arquivo CSV.\n"
                                "Adicionando a coluna 'periodicidade_dias' com valor padrão de 180 dias para carrinhos existentes.")

            temp_file = CSV_FILE + ".tmp"
            with open(CSV_FILE, "r", newline="", encoding="utf-8") as infile, \
                 open(temp_file, "w", newline="", encoding="utf-8") as outfile:
                reader = csv.reader(infile)
                writer = csv.writer(outfile)

                new_header = header + ["periodicidade_dias"]
                writer.writerow(new_header)

                for row in reader:
                    writer.writerow(row + ["180"])

            os.remove(CSV_FILE)
            os.rename(temp_file, CSV_FILE) 


# -------- Funções --------
def calcular_status(proxima_data):
    hoje = datetime.today().date()
    dias = (proxima_data - hoje).days
    if dias < 0:
        return "❌ VENCIDO", "#e74c3c"  
    elif dias <= 20:  
        return "VENCE EM BREVE", "#e74c3c" 
    elif dias <= 30: 
        return "ATENÇÃO: PRÓXIMO DO VENCIMENTO", "#f1c40f" 
    else:
        return "✅ EM DIA", "#2ecc71" 


def estilo_botao(botao, cor_bg, cor_hover):
    botao.config(
        bg=cor_bg,
        fg="white",
        relief="flat",
        font=("Segoe UI", 13, "bold"),
        width=18,
        height=2,
        bd=0,
        highlightthickness=0,
        cursor="hand2",
        activebackground=cor_hover,
    )

    def on_enter(e):
        botao.config(bg=cor_hover)

    def on_leave(e):
        botao.config(bg=cor_bg)

    botao.bind("<Enter>", on_enter)
    botao.bind("<Leave>", on_leave)


def buscar_carrinho():
    global carrinho_atual
    codigo = entrada.get().strip()
    carrinho_atual = None
    with open(CSV_FILE, newline="", encoding="utf-8") as csvfile:
        for linha in csv.DictReader(csvfile):
            if linha["codigo"] == codigo:
                data_ultima = datetime.strptime(
                    linha["data_ultima_limpeza"], "%Y-%m-%d"
                ).date()

                periodicidade_dias = int(linha.get("periodicidade_dias", 180)) 
                proxima = data_ultima + timedelta(days=periodicidade_dias)

                status, cor = calcular_status(proxima)
                resultado_frame.config(highlightbackground=cor, highlightthickness=3)
                resultado_label.config(
                    text=(
                        f"Carrinho: {codigo}\n\n"
                        f"🗓️ Última limpeza: {data_ultima.strftime('%d/%m/%Y')}\n"
                        f"📅 Periodicidade: {periodicidade_dias} dias\n" 
                        f"📅 Próxima limpeza: {proxima.strftime('%d/%m/%Y')}\n"
                        f"🔎 Status: {status}"
                    ),
                    fg="#2c3e50",
                )
                carrinho_atual = linha
                return
    resultado_frame.config(highlightbackground="#95a5a6", highlightthickness=3)
    resultado_label.config(text="❌ carrinho não encontrado", fg="#7f8c8d")


def cadastrar():
    codigo = entrada_codigo.get().strip()
    data = entrada_data.get().strip()
    historico = entrada_hist.get().strip()
    periodicidade_str = entrada_periodicidade.get().strip()

    try:
        data_formatada = datetime.strptime(data, "%d/%m/%Y").date()
    except ValueError:
        messagebox.showerror("Erro", "Data inválida! Use DD/MM/AAAA.")
        return
    if data_formatada > datetime.today().date():
        messagebox.showerror("Erro", "Data futura inválida.")
        return

    try:
        periodicidade_dias = int(periodicidade_str)
        if periodicidade_dias <= 0:
            messagebox.showerror("Erro", "Periodicidade deve ser um número inteiro positivo.")
            return
    except ValueError:
        messagebox.showerror("Erro", "Periodicidade inválida! Digite um número de dias.")
        return

    with open(CSV_FILE, "r", newline="", encoding="utf-8") as csvfile:
        for row in csv.DictReader(csvfile):
            if row["codigo"] == codigo:
                messagebox.showwarning("Aviso", f"carrinho '{codigo}' já existe.")
                return

    with open(CSV_FILE, "a", newline="", encoding="utf-8") as csvfile:
        csv.writer(csvfile).writerow([codigo, data_formatada.isoformat(), historico, str(periodicidade_dias)])

    messagebox.showinfo("Sucesso", "Carrinho cadastrado com sucesso!")
    janela_cadastro.destroy()


def abrir_cadastro():
    global janela_cadastro, entrada_codigo, entrada_data, entrada_hist, entrada_periodicidade
    janela_cadastro = tk.Toplevel(root)
    janela_cadastro.title("➕ Cadastrar Carrinho")
    janela_cadastro.configure(bg="#f4f6f9")
    janela_cadastro.geometry("500x500")
    janela_cadastro.transient(root)
    janela_cadastro.grab_set()

    tk.Label(
        janela_cadastro,
        text="Cadastro de Carrinho",
        font=("Segoe UI", 18, "bold"),
        bg="#f4f6f9",
        fg="#2c3e50",
    ).pack(pady=10)

    frame = tk.Frame(janela_cadastro, bg="white", relief="solid", bd=1)
    frame.pack(fill="both", expand=True, padx=20, pady=10)

    def campo(texto, valor_inicial=""):
        tk.Label(
            frame,
            text=texto,
            font=("Segoe UI", 13),
            bg="white",
            fg="#495057",
            anchor="w",
        ).pack(fill="x", padx=20, pady=(10, 0))
        entry = tk.Entry(frame, font=("Segoe UI", 13), relief="solid", bd=1)
        entry.pack(fill="x", padx=20, pady=5, ipady=5)
        entry.insert(0, valor_inicial)
        return entry

    entrada_codigo = campo("🔢 Código:")
    entrada_data = campo("📅 Data última limpeza (DD/MM/AAAA):")
    entrada_hist = campo("📝 Histórico:")
    entrada_periodicidade = campo("🗓️ Periodicidade de vencimento (dias):", "180")

    btn_salvar = tk.Button(janela_cadastro, text="Salvar Cadastro", command=cadastrar)
    estilo_botao(btn_salvar, "#27ae60", "#229954")
    btn_salvar.pack(pady=15, fill="x", padx=20)

    entrada_codigo.focus_set()
    janela_cadastro.wait_window(janela_cadastro)
    root.grab_release()


def editar_carrinho_unico():
    global carrinho_atual
    if not carrinho_atual:
        messagebox.showwarning("Aviso", "Consulte um carrinho para editar.")
        return

    def salvar_edicao():
        nova_data = entrada_data_edit.get().strip()
        novo_hist = entrada_hist_edit.get().strip()
        nova_periodicidade_str = entrada_periodicidade_edit.get().strip()

        try:
            data_formatada = datetime.strptime(nova_data, "%d/%m/%Y").date()
        except ValueError:
            messagebox.showerror("Erro", "Data inválida! Use DD/MM/AAAA.")
            return
        if data_formatada > datetime.today().date():
            messagebox.showerror("Erro", "Data futura inválida.")
            return

        try:
            nova_periodicidade_dias = int(nova_periodicidade_str)
            if nova_periodicidade_dias <= 0:
                messagebox.showerror("Erro", "Periodicidade deve ser um número inteiro positivo.")
                return
        except ValueError:
            messagebox.showerror("Erro", "Periodicidade inválida! Digite um número de dias.")
            return

        linhas = []
        with open(CSV_FILE, newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            header = reader.fieldnames

            for linha in reader:
                if linha["codigo"] == carrinho_atual["codigo"]:
                    linha["data_ultima_limpeza"] = data_formatada.isoformat()
                    linha["historico"] = novo_hist
                    linha["periodicidade_dias"] = str(nova_periodicidade_dias)
                linhas.append(linha)

        with open(CSV_FILE, "w", newline="", encoding="utf-8") as csvfile:
            escritor = csv.DictWriter(
                csvfile, fieldnames=header
            )
            escritor.writeheader()
            escritor.writerows(linhas)
        messagebox.showinfo("Sucesso", "Carrinho editado com sucesso!")
        janela_edicao.destroy()
        buscar_carrinho()
     
    janela_edicao = tk.Toplevel(root)
    janela_edicao.title(f"✏️ Editar Carrinho {carrinho_atual['codigo']}")
    janela_edicao.configure(bg="#f4f6f9")
    janela_edicao.geometry("500x450")
    janela_edicao.transient(root)
    janela_edicao.grab_set()

    tk.Label(
        janela_edicao,
        text=f"Editar Carrinho {carrinho_atual['codigo']}",
        font=("Segoe UI", 18, "bold"),
        bg="#f4f6f9",
        fg="#2c3e50",
    ).pack(pady=10)

    frame = tk.Frame(janela_edicao, bg="white", relief="solid", bd=1)
    frame.pack(fill="both", expand=True, padx=20, pady=10)

    tk.Label(
        frame,
        text="📅 Nova Data (DD/MM/AAAA):",
        font=("Segoe UI", 13),
        bg="white",
        fg="#495057",
    ).pack(pady=(10, 0))
    entrada_data_edit = tk.Entry(frame, font=("Segoe UI", 13))
    entrada_data_edit.pack(fill="x", padx=20, pady=5, ipady=5)
    entrada_data_edit.insert(
        0,
        datetime.strptime(carrinho_atual["data_ultima_limpeza"], "%Y-%m-%d").strftime(
            "%d/%m/%Y"
        ),
    )

    tk.Label(
        frame, text="📝 Histórico:", font=("Segoe UI", 13), bg="white", fg="#495057"
    ).pack(pady=(10, 0))
    entrada_hist_edit = tk.Entry(frame, font=("Segoe UI", 13))
    entrada_hist_edit.pack(fill="x", padx=20, pady=5, ipady=5)
    entrada_hist_edit.insert(0, carrinho_atual["historico"])

    tk.Label(
        frame, text="🗓️ Periodicidade (dias):", font=("Segoe UI", 13), bg="white", fg="#495057"
    ).pack(pady=(10, 0))
    entrada_periodicidade_edit = tk.Entry(frame, font=("Segoe UI", 13))
    entrada_periodicidade_edit.pack(fill="x", padx=20, pady=5, ipady=5)
    entrada_periodicidade_edit.insert(0, carrinho_atual.get("periodicidade_dias", "180"))

    btn = tk.Button(janela_edicao, text="Salvar Edição", command=salvar_edicao)
    estilo_botao(btn, "#f39c12", "#d68910")
    btn.pack(pady=15, fill="x", padx=20)

    entrada_data_edit.focus_set()
    janela_edicao.wait_window(janela_edicao)
    root.grab_release()

def abrir_edicao_em_lote():
    janela_selecao_carrinhos = tk.Toplevel(root)
    janela_selecao_carrinhos.title("✏️ Editar Vários Carrinhos")
    janela_selecao_carrinhos.configure(bg="#f4f6f9")
    janela_selecao_carrinhos.geometry("800x600")
    janela_selecao_carrinhos.transient(root)
    janela_selecao_carrinhos.grab_set()

    tk.Label(
        janela_selecao_carrinhos,
        text="Selecione os carrinhos para edição em lote",
        font=("Segoe UI", 18, "bold"),
        bg="#f4f6f9",
        fg="#2c3e50",
    ).pack(pady=10)

    frame_tree = tk.Frame(janela_selecao_carrinhos, bg="white", relief="solid", bd=1)
    frame_tree.pack(fill="both", expand=True, padx=20, pady=10)

    #Configurar o Treeview
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview.Heading", font=("Segoe UI", 12, "bold"), background="#34495e", foreground="white")
    style.configure("Treeview", font=("Segoe UI", 11), rowheight=25)
    style.map('Treeview', background=[('selected', '#2980b9')])

    tree = ttk.Treeview(frame_tree, columns=("codigo", "data_ultima_limpeza", "periodicidade_dias", "historico"), show="headings", selectmode="extended")
    tree.heading("codigo", text="Código")
    tree.heading("data_ultima_limpeza", text="Última Limpeza (DD/MM/AAAA)")
    tree.heading("periodicidade_dias", text="Periodicidade (dias)")
    tree.heading("historico", text="Histórico")

    tree.column("codigo", width=100, anchor="center")
    tree.column("data_ultima_limpeza", width=200, anchor="center")
    tree.column("periodicidade_dias", width=150, anchor="center")
    tree.column("historico", width=300, anchor="w")

    # Scrollbar
    scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    tree.pack(side="left", fill="both", expand=True)

    with open(CSV_FILE, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            tree.insert("", "end", iid=row["codigo"], values=(
                row["codigo"],
                datetime.strptime(row["data_ultima_limpeza"], "%Y-%m-%d").strftime("%d/%m/%Y"), 
                row.get("periodicidade_dias", "180"),
                row["historico"]
            ))

    def prosseguir_edicao_em_lote():
        selecionados = tree.selection() 
        if not selecionados:
            messagebox.showwarning("Aviso", "Nenhum carrinho selecionado para edição.")
            return

        janela_selecao_carrinhos.destroy()
        abrir_form_edicao_lote(list(selecionados)) 

    btn_selecionar = tk.Button(janela_selecao_carrinhos, text="Prosseguir para Edição", command=prosseguir_edicao_em_lote)
    estilo_botao(btn_selecionar, "#f39c12", "#d68910")
    btn_selecionar.pack(pady=15, fill="x", padx=20)

    janela_selecao_carrinhos.wait_window(janela_selecao_carrinhos)
    root.grab_release()

def abrir_form_edicao_lote(codigos_selecionados):
    janela_edicao_lote = tk.Toplevel(root)
    janela_edicao_lote.title(f"✏️ Editar {len(codigos_selecionados)} Carrinhos")
    janela_edicao_lote.configure(bg="#f4f6f9")
    janela_edicao_lote.geometry("500x650") 
    janela_edicao_lote.transient(root)
    janela_edicao_lote.grab_set()

    tk.Label(
        janela_edicao_lote,
        text=f"Editar em Lote ({len(codigos_selecionados)} carrinhos)",
        font=("Segoe UI", 18, "bold"),
        bg="#f4f6f9",
        fg="#2c3e50",
    ).pack(pady=10)

    tk.Label(
        janela_edicao_lote,
        text="Preencha APENAS os campos que deseja alterar para todos os carrinhos selecionados.\nDeixe em branco para MANTER o valor atual.",
        font=("Segoe UI", 10, "italic"),
        bg="#f4f6f9",
        fg="#7f8c8d",
        justify="center"
    ).pack(pady=(0, 10))

    tk.Label(
        janela_edicao_lote,
        text="Carrinhos Selecionados:",
        font=("Segoe UI", 12, "bold"),
        bg="#f4f6f9",
        fg="#2c3e50",
    ).pack(pady=(10, 0))

    text_carrinhos = tk.Text(janela_edicao_lote, height=5, width=40, font=("Segoe UI", 10), wrap="word", relief="solid", bd=1)
    text_carrinhos.pack(pady=5, padx=20, fill="x")
    for codigo in codigos_selecionados:
        text_carrinhos.insert(tk.END, f"- {codigo}\n")
    text_carrinhos.config(state=tk.DISABLED)

    scrollbar_text = tk.Scrollbar(janela_edicao_lote, command=text_carrinhos.yview)
    scrollbar_text.pack(side="right", fill="y")
    text_carrinhos.config(yscrollcommand=scrollbar_text.set)

    frame = tk.Frame(janela_edicao_lote, bg="white", relief="solid", bd=1)
    frame.pack(fill="both", expand=True, padx=20, pady=10)

    tk.Label(
        frame,
        text="📅 Nova Data de Limpeza (DD/MM/AAAA):",
        font=("Segoe UI", 13),
        bg="white",
        fg="#495057",
    ).pack(pady=(10, 0))
    entrada_nova_data = tk.Entry(frame, font=("Segoe UI", 13))
    entrada_nova_data.pack(fill="x", padx=20, pady=5, ipady=5)

    tk.Label(
        frame, text="📝 Novo Histórico (será SOBRESCRITO):", font=("Segoe UI", 13), bg="white", fg="#495057"
    ).pack(pady=(10, 0))
    entrada_novo_historico = tk.Entry(frame, font=("Segoe UI", 13))
    entrada_novo_historico.pack(fill="x", padx=20, pady=5, ipady=5)

    tk.Label(
        frame, text="🗓️ Nova Periodicidade (dias):", font=("Segoe UI", 13), bg="white", fg="#495057"
    ).pack(pady=(10, 0))
    entrada_nova_periodicidade = tk.Entry(frame, font=("Segoe UI", 13))
    entrada_nova_periodicidade.pack(fill="x", padx=20, pady=5, ipady=5)

    def aplicar_edicoes_lote():
        nova_data_str = entrada_nova_data.get().strip()
        novo_historico = entrada_novo_historico.get().strip()
        nova_periodicidade_str = entrada_nova_periodicidade.get().strip()

        data_formatada = None
        if nova_data_str:
            try:
                data_formatada = datetime.strptime(nova_data_str, "%d/%m/%Y").date()
                if data_formatada > datetime.today().date():
                    messagebox.showerror("Erro", "Data futura inválida.")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Data inválida! Use DD/MM/AAAA.")
                return
        
        nova_periodicidade_dias = None
        if nova_periodicidade_str:
            try:
                nova_periodicidade_dias = int(nova_periodicidade_str)
                if nova_periodicidade_dias <= 0:
                    messagebox.showerror("Erro", "Periodicidade deve ser um número inteiro positivo.")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Periodicidade inválida! Digite um número de dias.")
                return

        if not nova_data_str and not novo_historico and not nova_periodicidade_str:
            messagebox.showwarning("Aviso", "Nenhum campo foi preenchido para edição. Nenhuma alteração será aplicada.")
            return

        linhas_atualizadas = []
        with open(CSV_FILE, newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            header = reader.fieldnames
            for linha in reader:
                if linha["codigo"] in codigos_selecionados:
                    if data_formatada:
                        linha["data_ultima_limpeza"] = data_formatada.isoformat()
                    if novo_historico:
                        linha["historico"] = novo_historico
                    if nova_periodicidade_dias:
                        linha["periodicidade_dias"] = str(nova_periodicidade_dias)
                linhas_atualizadas.append(linha)
        
        with open(CSV_FILE, "w", newline="", encoding="utf-8") as csvfile:
            escritor = csv.DictWriter(csvfile, fieldnames=header)
            escritor.writeheader()
            escritor.writerows(linhas_atualizadas)
        
        messagebox.showinfo("Sucesso", f"{len(codigos_selecionados)} carrinhos foram atualizados com sucesso!")
        janela_edicao_lote.destroy()
        if carrinho_atual and carrinho_atual["codigo"] in codigos_selecionados:
            buscar_carrinho() 

    btn_aplicar = tk.Button(janela_edicao_lote, text="Aplicar Edições em Lote", command=aplicar_edicoes_lote)
    estilo_botao(btn_aplicar, "#27ae60", "#229954")
    btn_aplicar.pack(pady=15, fill="x", padx=20)

    janela_edicao_lote.wait_window(janela_edicao_lote)
    root.grab_release()


def gerar_relatorio():
    hoje = datetime.today().date()
    if hoje.month == 12:
        primeiro_dia_proximo_mes = hoje.replace(year=hoje.year + 1, month=1, day=1)
    else:
        primeiro_dia_proximo_mes = hoje.replace(month=hoje.month + 1, day=1)

    carrinhos_proximo_mes = []
    with open(CSV_FILE, newline="", encoding="utf-8") as csvfile:
        for linha in csv.DictReader(csvfile):
            data_ultima = datetime.strptime(
                linha["data_ultima_limpeza"], "%Y-%m-%d"
            ).date()

            periodicidade_dias = int(linha.get("periodicidade_dias", 180))
            proxima = data_ultima + timedelta(days=periodicidade_dias)

            if (proxima.year == primeiro_dia_proximo_mes.year and proxima.month == primeiro_dia_proximo_mes.month) or \
               (proxima < hoje):
                carrinhos_proximo_mes.append(
                    {
                        "codigo": linha["codigo"],
                        "ultima": data_ultima, 
                        "proxima": proxima, 
                        "periodicidade": periodicidade_dias,
                        "historico": linha["historico"],
                        "status_vencimento": "VENCIDO" if proxima < hoje else "A VENCER"
                    }
                )

    carrinhos_proximo_mes.sort(key=lambda x: x['proxima']) 
    if not os.path.exists("relatorios"):
        os.makedirs("relatorios")

    file_path_excel = f"relatorios/relatorio_carrinhos_{primeiro_dia_proximo_mes.month:02d}_{primeiro_dia_proximo_mes.year}.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Carrinhos"

    header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
    title_font = Font(name='Segoe UI', size=16, bold=True, color='2C3E50')
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    fill_header = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    fill_vencido = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    fill_avencer = PatternFill(start_color='F1C40F', end_color='F1C40F', fill_type='solid') 
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')

    sheet.merge_cells('A1:F1')
    title_cell = sheet['A1']
    title_cell.value = "RELATÓRIO DE LIMPEZA DE CARRINHOS"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid') 

    sheet.merge_cells('A2:F2')
    subtitle_cell = sheet['A2']
    subtitle_cell.value = f"Referente ao mês {primeiro_dia_proximo_mes.month:02d}/{primeiro_dia_proximo_mes.year} e Vencidos"
    subtitle_cell.font = Font(name='Segoe UI', size=12, bold=True, color='7F8C8D')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    subtitle_cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')

    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 20

    headers = ["CÓDIGO", "ÚLTIMA LIMPEZA", "PRÓXIMA LIMPEZA", "PERIOD. (dias)", "STATUS", "HISTÓRICO"]
    sheet.append([]) 
    sheet.append(headers) 

    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_idx) 
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_style
        cell.alignment = alignment_center

    row_num = 5
    if carrinhos_proximo_mes:
        for c in carrinhos_proximo_mes:
            sheet.cell(row=row_num, column=1, value=c['codigo']).alignment = alignment_center
            sheet.cell(row=row_num, column=2, value=c['ultima']).number_format = 'DD/MM/YYYY'
            sheet.cell(row=row_num, column=3, value=c['proxima']).number_format = 'DD/MM/YYYY'
            sheet.cell(row=row_num, column=4, value=c['periodicidade']).alignment = alignment_center
            sheet.cell(row=row_num, column=5, value=c['status_vencimento']).alignment = alignment_center
            status_cell = sheet.cell(row=row_num, column=5)
            if c['status_vencimento'] == "VENCIDO":
                status_cell.fill = fill_vencido
                status_cell.font = Font(color='FFFFFF', bold=True)
            elif c['status_vencimento'] == "A VENCER":
                status_cell.fill = fill_avencer
                status_cell.font = Font(color='000000', bold=True)
            sheet.cell(row=row_num, column=6, value=c['historico']).alignment = alignment_left

            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_num, column=col_idx)
                cell.border = border_style

            row_num += 1
    else:
        sheet.merge_cells(f'A5:F5')
        sheet['A5'].value = "Não há carrinhos para relatar neste período."
        sheet['A5'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A5'].font = Font(name='Segoe UI', size=12, italic=True, color='7F8C8D')

        for col_idx in range(1, len(headers) + 1):
            sheet.cell(row=5, column=col_idx).border = border_style


    column_widths = [12, 18, 18, 15, 15, 40] 
    for i, width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i+1)].width = width

    try:
        workbook.save(file_path_excel)
        messagebox.showinfo("Relatório", f"Relatório gerado com sucesso!\n\nArquivo Excel: '{file_path_excel}'")
    except Exception as e:
        messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o arquivo Excel. Erro: {e}")


def verificar_senha(callback_funcao):
    """
    Abre uma janela para solicitar a senha. Se a senha estiver correta,
    chama a 'callback_funcao'.
    """
    janela_senha = tk.Toplevel(root)
    janela_senha.title("Autenticação")
    janela_senha.geometry("300x180")
    janela_senha.configure(bg="#f4f6f9")
    janela_senha.resizable(False, False)
    janela_senha.transient(root)
    janela_senha.grab_set() 

    def tentar_login():
        senha_digitada = entrada_senha.get()
        if senha_digitada == SENHA_MESTRE:
            janela_senha.destroy()
            callback_funcao()
        else:
            messagebox.showerror("Erro de Senha", "Senha incorreta!")
            entrada_senha.delete(0, tk.END) 
            entrada_senha.focus_set()

    tk.Label(janela_senha, text="Digite a senha para acessar:",
             font=("Segoe UI", 12), bg="#f4f6f9", fg="#2c3e50").pack(pady=15)

    entrada_senha = tk.Entry(janela_senha, show="*", font=("Segoe UI", 14), relief="solid", bd=1)
    entrada_senha.pack(pady=5, padx=20, fill="x", ipady=5)
    entrada_senha.bind("<Return>", lambda e: tentar_login()) 
    entrada_senha.focus_set()

    btn_login = tk.Button(janela_senha, text="Entrar", command=tentar_login)
    estilo_botao(btn_login, "#3498db", "#2980b9") 
    btn_login.config(width=10, height=1) 
    btn_login.pack(pady=15)

    janela_senha.update_idletasks()
    x = root.winfo_x() + (root.winfo_width() // 2) - (janela_senha.winfo_width() // 2)
    y = root.winfo_y() + (root.winfo_height() // 2) - (janela_senha.winfo_height() // 2)
    janela_senha.geometry(f"+{x}+{y}")


root = tk.Tk()
root.title(" Controle de Limpeza de Carrinhos")
root.state("zoomed")
root.configure(bg="#f4f6f9")

tk.Label(
    root,
    text=" Controle de Limpeza de Carrinhos",
    font=("Segoe UI", 32, "bold"),
    bg="#f4f6f9",
    fg="#2c3e50",
).pack(pady=20)

main_frame = tk.Frame(root, bg="white", relief="groove", bd=2)
main_frame.pack(fill="both", expand=True, padx=80, pady=30)

entrada = tk.Entry(
    main_frame, font=("Segoe UI", 20), relief="solid", justify="center", bd=2
)
entrada.insert(0, "Digite o código do carrinho...")
entrada.config(fg="gray")
entrada.bind(
    "<FocusIn>", lambda e: (entrada.delete(0, "end"), entrada.config(fg="black"))
)
entrada.bind("<Return>", lambda e: buscar_carrinho())
entrada.pack(pady=30, padx=200, ipady=10, fill="x")

btn_frame = tk.Frame(main_frame, bg="white")
btn_frame.pack(pady=20)

btn_editar = tk.Menubutton(btn_frame, text="✏️ Editar", relief="flat", font=("Segoe UI", 13, "bold"),
                           bg="#f39c12", fg="white", width=18, height=2, bd=0,
                           highlightthickness=0, cursor="hand2", activebackground="#d68910")
estilo_botao(btn_editar, "#f39c12", "#d68910") 
btn_editar.grid(row=0, column=2, padx=15) 

menu_edicao = tk.Menu(btn_editar, tearoff=0, font=("Segoe UI", 12))
menu_edicao.add_command(label="Editar Carrinho Selecionado", command=lambda: verificar_senha(editar_carrinho_unico))
menu_edicao.add_command(label="Editar em Lote", command=lambda: verificar_senha(abrir_edicao_em_lote))
btn_editar["menu"] = menu_edicao

btns_esquerda = [
    ("🔍 Consultar", buscar_carrinho, "#34495e", "#2c3e50"),
    ("➕ Novo Cadastro", abrir_cadastro, "#27ae60", "#229954"),
]

btns_direita = [
    ("📄 Relatório", gerar_relatorio, "#2980b9", "#2471a3"),
]

for i, (txt, cmd, c1, c2) in enumerate(btns_esquerda):
    b = tk.Button(btn_frame, text=txt, command=cmd)
    estilo_botao(b, c1, c2)
    b.grid(row=0, column=i, padx=15)

for i, (txt, cmd, c1, c2) in enumerate(btns_direita):
    b = tk.Button(btn_frame, text=txt, command=cmd)
    estilo_botao(b, c1, c2)
    b.grid(row=0, column=i + 3, padx=15) 

resultado_frame = tk.Frame(
    main_frame,
    bg="white",
    relief="solid",
    bd=2,
    highlightthickness=3,
    highlightbackground="#95a5a6",
)
resultado_frame.pack(fill="both", expand=True, padx=60, pady=30)
resultado_label = tk.Label(
    resultado_frame,
    text="",
    font=("Segoe UI", 18),
    bg="white",
    fg="#2c3e50",
    justify="left",
    anchor="nw",
)
resultado_label.pack(padx=30, pady=30, anchor="w")

root.mainloop()